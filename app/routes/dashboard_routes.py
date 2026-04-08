from flask import Blueprint, render_template, session, redirect, send_file, request, flash
from ..models.db import get_connection
from ..services.risk_service import predict_risk
from ..services.timetable_service import get_current_subject, get_all_timetable, save_timetable
from ..services.face_service import capture_face_image
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
from datetime import datetime
import calendar
from openpyxl.styles import PatternFill

dashboard_bp = Blueprint("dashboard", __name__)

@dashboard_bp.route("/dashboard")
def dashboard():
    if "role" not in session:
        return redirect("/")
    role = session.get("role")
    if role == "student":
        return redirect("/student_dashboard")

    conn = get_connection()
    c = conn.cursor()

    timetable = get_all_timetable()
    if role == "admin":
        total_students = c.execute("SELECT COUNT(*) FROM students").fetchone()[0]
        # only count assigned CSE AIML faculty (non-empty subject, valid name)
        total_faculty = c.execute(
            "SELECT COUNT(*) FROM users u JOIN faculty f ON u.id=f.user_id WHERE u.role='faculty' AND f.department='CSE AIML' AND f.assigned_subject IS NOT NULL AND TRIM(f.assigned_subject)<>'' AND f.name IS NOT NULL AND TRIM(LOWER(f.name))<>'none'"
        ).fetchone()[0]
        total_attendance = c.execute("SELECT COUNT(*) FROM attendance").fetchone()[0]

        student_records = c.execute(
            "SELECT s.id, s.name, s.roll, COUNT(a.id) FROM students s LEFT JOIN attendance a ON s.id=a.student_id GROUP BY s.id"
        ).fetchall()
        faculty_records = c.execute(
            "SELECT u.username, f.name, f.department, f.assigned_subject FROM users u JOIN faculty f ON u.id=f.user_id WHERE u.role='faculty' AND f.department='CSE AIML' AND f.assigned_subject IS NOT NULL AND TRIM(f.assigned_subject)<>'' AND f.name IS NOT NULL AND TRIM(LOWER(f.name))<>'none'"
        ).fetchall()
        all_attendance = c.execute(
            "SELECT a.date, a.time, a.subject, s.name FROM attendance a LEFT JOIN students s ON a.student_id=s.id ORDER BY a.date DESC, a.time DESC LIMIT 20"
        ).fetchall()

        conn.close()
        stats = {
            "students": total_students,
            "faculty": total_faculty,
            "attendance": total_attendance,
        }
        deleted_message = request.args.get('deleted')
        return render_template(
            "admin_dashboard.html",
            stats=stats,
            student_records=student_records,
            faculty_records=faculty_records,
            all_attendance=all_attendance,
            timetable=timetable,
            deleted_message=deleted_message,
        )

    if role == "faculty":
        faculty_subject_row = c.execute(
            "SELECT f.assigned_subject FROM faculty f JOIN users u ON u.id=f.user_id WHERE u.id=?", (session.get('user_id'),)
        ).fetchone()
        assigned_subject = faculty_subject_row[0] if faculty_subject_row and faculty_subject_row[0] else None

        if assigned_subject:
            total_students = c.execute("SELECT COUNT(*) FROM students").fetchone()[0]
            total_attendance = c.execute("SELECT COUNT(*) FROM attendance WHERE subject=?", (assigned_subject,)).fetchone()[0]
            total_subjects = 1
            recent = c.execute(
                """
                SELECT a.date, a.time, a.subject, s.name
                FROM attendance a
                LEFT JOIN students s ON a.student_id = s.id
                WHERE a.subject=?
                ORDER BY a.date DESC
                LIMIT 5
                """,
                (assigned_subject,),
            ).fetchall()
        else:
            total_students = c.execute("SELECT COUNT(*) FROM students").fetchone()[0]
            total_attendance = c.execute("SELECT COUNT(*) FROM attendance").fetchone()[0]
            total_subjects = c.execute("SELECT COUNT(DISTINCT subject) FROM attendance").fetchone()[0]
            recent = c.execute(
                """
                SELECT a.date, a.time, a.subject, s.name
                FROM attendance a
                LEFT JOIN students s ON a.student_id = s.id
                ORDER BY a.date DESC
                LIMIT 5
                """
            ).fetchall()

        conn.close()
        stats = {
            "students": total_students,
            "attendance": total_attendance,
            "subjects": total_subjects,
            "average": round(total_attendance / total_students, 1) if total_students else 0,
        }
        return render_template("dashboard.html", stats=stats, recent=recent, timetable=timetable, role=role, assigned_subject=assigned_subject)

# Add admin helper to delete all malformed/unassigned faculty entries where name is none/blank
@dashboard_bp.route('/delete_faculty_none', methods=['POST'])
def delete_faculty_none():
    if session.get('role') != 'admin':
        return redirect('/dashboard')
    conn = get_connection()
    c = conn.cursor()
    rows = c.execute("SELECT id, user_id FROM faculty WHERE LOWER(TRIM(name))='none' OR name IS NULL OR TRIM(name)='' ").fetchall()
    for fid, uid in rows:
        c.execute('DELETE FROM faculty WHERE id=?', (fid,))
        if uid:
            c.execute('DELETE FROM users WHERE id=?', (uid,))
    conn.commit()
    conn.close()
    return redirect('/dashboard?deleted=1')


@dashboard_bp.route('/edit_students', methods=['GET', 'POST'])
def edit_students():
    if session.get('role') != 'admin':
        return redirect('/dashboard')

    conn = get_connection()
    c = conn.cursor()
    msg = None

    if request.method == 'POST':
        action = request.form.get('action')
        student_id = request.form.get('student_id')
        if student_id:
            if action == 'update':
                name = request.form.get('name')
                roll = request.form.get('roll')
                c.execute('UPDATE students SET name=?, roll=? WHERE id=?', (name, roll, student_id))
                msg = 'Student record updated successfully.'
            elif action == 'delete':
                student = c.execute('SELECT user_id, image FROM students WHERE id=?', (student_id,)).fetchone()
                if student:
                    user_id, image = student
                    c.execute('DELETE FROM attendance WHERE student_id=?', (student_id,))
                    c.execute('DELETE FROM students WHERE id=?', (student_id,))
                    if user_id:
                        c.execute('DELETE FROM users WHERE id=?', (user_id,))
                    if image and os.path.exists(image):
                        try:
                            os.remove(image)
                        except Exception:
                            pass
                    msg = 'Student deleted successfully.'
    rows = c.execute('SELECT id, name, roll, image FROM students ORDER BY id').fetchall()
    conn.commit()
    conn.close()
    return render_template('edit_students.html', rows=rows, msg=msg)


@dashboard_bp.route('/delete_student/<int:student_id>', methods=['POST'])
def delete_student(student_id):
    if session.get('role') != 'admin':
        return redirect('/dashboard')
    conn = get_connection()
    c = conn.cursor()
    student = c.execute('SELECT user_id, image FROM students WHERE id=?', (student_id,)).fetchone()
    if student:
        user_id, image = student
        c.execute('DELETE FROM attendance WHERE student_id=?', (student_id,))
        c.execute('DELETE FROM students WHERE id=?', (student_id,))
        if user_id:
            c.execute('DELETE FROM users WHERE id=?', (user_id,))
        if image and os.path.exists(image):
            try:
                os.remove(image)
            except Exception:
                pass
    conn.commit()
    conn.close()
    return redirect('/dashboard')


@dashboard_bp.route('/edit_faculty', methods=['GET', 'POST'])
def edit_faculty():
    if session.get('role') != 'admin':
        return redirect('/dashboard')

    conn = get_connection()
    c = conn.cursor()
    msg = None

    if request.method == 'POST':
        action = request.form.get('action')
        faculty_id = request.form.get('faculty_id')
        if action == 'update' and faculty_id:
            name = request.form.get('name')
            department = request.form.get('department')
            assigned_subject = request.form.get('assigned_subject', '').strip()
            c.execute("UPDATE faculty SET name=?, department=?, assigned_subject=? WHERE id=?", (name, department, assigned_subject, faculty_id))
            msg = "Faculty updated successfully."
        elif action == 'delete' and faculty_id:
            fac = c.execute("SELECT user_id, name FROM faculty WHERE id=?", (faculty_id,)).fetchone()
            if fac:
                user_id, name = fac
                if not name or not name.strip():
                    c.execute("DELETE FROM faculty WHERE id=?", (faculty_id,))
                    c.execute("DELETE FROM users WHERE id=?", (user_id,))
                    msg = "Faculty with no name/records deleted successfully."
                else:
                    msg = "Cannot delete faculty with existing record name. Clear name first to delete."
        conn.commit()
        rows = c.execute(
            "SELECT f.id, u.username, f.name, f.department, f.assigned_subject FROM faculty f JOIN users u ON u.id=f.user_id ORDER BY f.id"
        ).fetchall()
        conn.close()
        return render_template('edit_faculty.html', rows=rows, msg=msg)

    rows = c.execute(
        "SELECT f.id, u.username, f.name, f.department, f.assigned_subject FROM faculty f JOIN users u ON u.id=f.user_id ORDER BY f.id"
    ).fetchall()
    conn.close()
    return render_template('edit_faculty.html', rows=rows, msg=msg)


@dashboard_bp.route("/cse_faculty")
def cse_faculty():
    if session.get('role') not in ['faculty', 'admin']:
        return redirect('/')
    conn = get_connection()
    c = conn.cursor()
    faculty = c.execute(
        "SELECT u.username, f.name, f.department FROM faculty f JOIN users u ON u.id=f.user_id WHERE f.department='CSE AIML'"
    ).fetchall()
    conn.close()
    return render_template('cse_faculty.html', faculty=faculty)


@dashboard_bp.route('/edit_attendance', methods=['GET', 'POST'])
def edit_attendance():
    if session.get('role') not in ['faculty', 'admin']:
        return redirect('/')
    conn = get_connection()
    c = conn.cursor()
    if request.method == 'POST':
        action = request.form.get('action')
        att_id = request.form.get('attendance_id')
        if action == 'update' and att_id:
            subject = request.form.get('subject')
            date = request.form.get('date')
            time = request.form.get('time')
            c.execute("UPDATE attendance SET subject=?, date=?, time=? WHERE id=?", (subject, date, time, att_id))
        elif action == 'delete' and att_id:
            c.execute("DELETE FROM attendance WHERE id=?", (att_id,))
        conn.commit()

    rows = c.execute(
        "SELECT a.id, s.name, s.roll, a.subject, a.date, a.time FROM attendance a LEFT JOIN students s ON s.id=a.student_id ORDER BY a.date DESC, a.time DESC"
    ).fetchall()
    conn.close()
    return render_template('edit_attendance.html', rows=rows)


@dashboard_bp.route("/upload_attendance", methods=["POST"])
def upload_attendance():
    if session.get('role') not in ['faculty', 'admin']:
        return redirect('/')

    file = request.files.get('attendance_file')
    if not file:
        return redirect('/dashboard')

    wb = openpyxl.load_workbook(file)
    ws = wb.active
    conn = get_connection()
    c = conn.cursor()
    added = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 4:
            continue
        date, time, subject, roll = row[:4]
        if not (date and time and subject and roll):
            continue

        student = c.execute("SELECT id FROM students WHERE roll=?", (roll,)).fetchone()
        if not student:
            continue
        student_id = student[0]
        c.execute("SELECT id FROM attendance WHERE student_id=? AND subject=? AND date=? AND time=?", (student_id, subject, date, time))
        if not c.fetchone():
            c.execute("INSERT INTO attendance(student_id, subject, date, time) VALUES (?, ?, ?, ?)", (student_id, subject, date, time))
            added += 1
    conn.commit()
    conn.close()
    return redirect('/dashboard')


@dashboard_bp.route("/student_dashboard")
def student_dashboard():
    if session.get("role") != "student":
        return redirect("/dashboard")

    user_id = session["user_id"]
    conn = get_connection()
    c = conn.cursor()

    student = c.execute(
        "SELECT id FROM students WHERE user_id=?",
        (user_id,)
    ).fetchone()

    student_id = student[0]

    data = c.execute("""
        SELECT subject, COUNT(id)
        FROM attendance
        WHERE student_id=?
        GROUP BY subject
    """, (student_id,)).fetchall()

    result = []

    for subject, attended in data:
        total = c.execute("""
            SELECT COUNT(DISTINCT date)
            FROM attendance
            WHERE subject=?
        """, (subject,)).fetchone()[0]

        risk, percent = predict_risk(attended, total)
        result.append((subject, attended, total, round(percent,2), risk))

    records = c.execute(
        """
        SELECT subject, date, time
        FROM attendance
        WHERE student_id=?
        ORDER BY date DESC, time DESC
        """,
        (student_id,)
    ).fetchall()

    conn.close()

    timetable = get_all_timetable()
    return render_template("student_dashboard.html", result=result, records=records, timetable=timetable)


@dashboard_bp.route('/download_student_monthly_attendance')
def download_student_monthly_attendance():
    if session.get('role') != 'student':
        return redirect('/')

    month_year = request.args.get('month_year')
    if not month_year:
        flash('Please select month and year to download monthly attendance report.')
        return redirect('/student_dashboard')

    try:
        report_date = datetime.strptime(month_year, '%Y-%m')
    except ValueError:
        flash('Invalid month format. Use YYYY-MM.')
        return redirect('/student_dashboard')

    year = report_date.year
    month = report_date.month
    num_days = calendar.monthrange(year, month)[1]

    user_id = session['user_id']
    conn = get_connection()
    c = conn.cursor()
    student = c.execute("SELECT id, name, roll FROM students WHERE user_id=?", (user_id,)).fetchone()
    if not student:
        conn.close()
        flash('Student not found.')
        return redirect('/student_dashboard')

    student_id, student_name, roll = student

    rows = c.execute(
        "SELECT subject, date FROM attendance WHERE student_id=?",
        (student_id,)
    ).fetchall()

    # include all subjects from attendance and timetable for full report
    subject_days = {}
    attendance_subjects = set()
    for subject, date_str in rows:
        if subject:
            attendance_subjects.add(subject.strip())
        if not subject or not date_str:
            continue
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            continue
        if dt.year == year and dt.month == month:
            subject_days.setdefault(subject.strip(), set()).add(dt.day)

    # include all subjects from timetable
    timetable_subjects = {entry['subject'].strip() for entry in get_all_timetable() if entry.get('subject')}
    all_subjects = sorted({s for s in attendance_subjects if s} | timetable_subjects)
    if not all_subjects:
        all_subjects = ['No Records']
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Monthly Attendance'

    ws.append(['Student Name:', student_name])
    ws.append(['Month/Year:', report_date.strftime('%B %Y')])
    ws.append([])

    header = ['Subject Name'] + [str(d) for d in range(1, num_days + 1)]
    ws.append(header)

    today = datetime.today().date()
    for subject in all_subjects:
        present_days = subject_days.get(subject, set())
        row = [subject]
        for day in range(1, num_days + 1):
            # future day in selected month -> leave empty
            current_date = datetime(year=year, month=month, day=day).date()
            if day in present_days:
                row.append('Present')
            elif current_date > today:
                row.append('')
            else:
                row.append('Absent')
        ws.append(row)
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=ws.max_row, column=col_index)

            if value == "Present":
               cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

            elif value == "Absent":
                 cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Set width for readability
    ws.column_dimensions[get_column_letter(1)].width = 26
    for col in range(2, num_days + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    month_name = report_date.strftime('%B')
    filename = f"attendance_all_{month_name}_{year}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@dashboard_bp.route('/timetable')
def timetable():
    if session.get('role') not in ['student', 'faculty', 'admin']:
        return redirect('/')
    entries = get_all_timetable()
    return render_template('timetable.html', entries=entries, role=session.get('role'))


@dashboard_bp.route('/edit_timetable', methods=['GET', 'POST'])
def edit_timetable():
    if session.get('role') != 'admin':
        return redirect('/dashboard')

    msg = None
    if request.method == 'POST':
        days = request.form.getlist('day')
        starts = request.form.getlist('start')
        ends = request.form.getlist('end')
        subjects = request.form.getlist('subject')
        entries = []
        for d, s, e, sub in zip(days, starts, ends, subjects):
            if d.strip() and s.strip() and e.strip() and sub.strip():
                entries.append({'day': d.strip(), 'start': s.strip(), 'end': e.strip(), 'subject': sub.strip()})
        save_timetable(entries)
        msg = 'Timetable updated successfully.'

    entries = get_all_timetable()
    return render_template('edit_timetable.html', entries=entries, msg=msg)


@dashboard_bp.route('/admin_capture_students')
def admin_capture_students():
    if session.get('role') != 'admin':
        return redirect('/dashboard')
    conn = get_connection()
    c = conn.cursor()
    students = c.execute('SELECT id, name, roll, image FROM students').fetchall()
    conn.close()
    return render_template('admin_capture_students.html', students=students)


@dashboard_bp.route('/capture_student_face/<int:student_id>', methods=['POST'])
def capture_student_face(student_id):
    if session.get('role') != 'admin':
        return redirect('/dashboard')

    conn = get_connection()
    c = conn.cursor()
    student = c.execute('SELECT name, roll FROM students WHERE id=?', (student_id,)).fetchone()
    if not student:
        conn.close()
        return redirect('/admin_capture_students')

    name, roll = student
    os.makedirs('static/faces', exist_ok=True)
    image_path = f'static/faces/{roll}_{student_id}.jpg'
    c.execute('UPDATE students SET image=? WHERE id=?', (image_path, student_id))
    conn.commit()
    conn.close()

    try:
        capture_face_image(image_path)
    except Exception:
        pass
    return redirect('/admin_capture_students')


@dashboard_bp.route('/download_attendance')
def download_attendance():
    if session.get('role') not in ['faculty', 'admin']:
        return redirect('/')

    subject = get_current_subject()
    conn = get_connection()
    c = conn.cursor()

    if subject:
        rows = c.execute(
            "SELECT a.subject, s.roll, s.name, a.date, a.time FROM attendance a LEFT JOIN students s ON a.student_id=s.id WHERE a.subject=? ORDER BY a.date DESC, a.time DESC",
            (subject,)
        ).fetchall()
    else:
        rows = c.execute(
            "SELECT a.subject, s.roll, s.name, a.date, a.time FROM attendance a LEFT JOIN students s ON a.student_id=s.id ORDER BY a.date DESC, a.time DESC"
        ).fetchall()

    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'
    ws.append(['Subject', 'Roll', 'Name', 'Date', 'Time'])
    for r in rows:
        ws.append(r)

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_{subject or 'all'}_.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@dashboard_bp.route('/download_subject_monthly_attendance')
def download_subject_monthly_attendance():
    if session.get('role') not in ['faculty', 'admin']:
        return redirect('/dashboard')

    subject = request.args.get('subject', '').strip()
    month_year = request.args.get('month_year', '').strip()
    if not month_year:
        flash('Please select month and year to download monthly report.')
        return redirect('/dashboard')

    if not subject:
        # for faculty, default to assigned subject if available
        if session.get('role') == 'faculty':
            subject = get_current_subject() or ''
    if not subject:
        flash('Please provide subject name to download report.')
        return redirect('/dashboard')

    try:
        report_date = datetime.strptime(month_year, '%Y-%m')
    except ValueError:
        flash('Invalid month format. Use YYYY-MM.')
        return redirect('/dashboard')

    year = report_date.year
    month = report_date.month
    num_days = calendar.monthrange(year, month)[1]

    conn = get_connection()
    c = conn.cursor()

    students = c.execute('SELECT id, roll, name FROM students ORDER BY roll').fetchall()
    attendance_rows = c.execute(
        'SELECT student_id, date FROM attendance WHERE subject=?',
        (subject,)
    ).fetchall()
    conn.close()

    attendance_by_student = {}
    for student_id, date_str in attendance_rows:
        if not student_id or not date_str:
            continue
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            continue
        if dt.year == year and dt.month == month:
            attendance_by_student.setdefault(student_id, set()).add(dt.day)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Subject Monthly Attendance'
    ws.append(['Subject:', subject])
    ws.append(['Month/Year:', report_date.strftime('%B %Y')])
    ws.append([])

    header = ['Roll No.', 'Student Name'] + [str(d) for d in range(1, num_days + 1)]
    ws.append(header)

    today = datetime.today().date()
    for student_id, roll, student_name in students:
        row = [roll or f'Roll {student_id}', student_name or f'Student {student_id}']
        present_days = attendance_by_student.get(student_id, set())
        for day in range(1, num_days + 1):
            current_date = datetime(year=year, month=month, day=day).date()
            if day in present_days:
                row.append('Present')
            elif current_date > today:
                row.append('')
            else:
                row.append('Absent')
        ws.append(row)
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=ws.max_row, column=col_index)

            if value == "Present":
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

            elif value == "Absent":
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    ws.column_dimensions[get_column_letter(1)].width = 26
    for col in range(2, num_days + 2):
        ws.column_dimensions[get_column_letter(col)].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='attendance_all.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@dashboard_bp.route('/download_student_attendance')
def download_student_attendance():
    if session.get('role') != 'student':
        return redirect('/')

    user_id = session['user_id']
    conn = get_connection()
    c = conn.cursor()
    student = c.execute("SELECT id, name, roll FROM students WHERE user_id=?", (user_id,)).fetchone()
    if not student:
        conn.close()
        return redirect('/student_dashboard')
    student_id, student_name, roll = student

    rows = c.execute(
        "SELECT subject, date, time FROM attendance WHERE student_id=? ORDER BY date DESC, time DESC",
        (student_id,)
    ).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'My Attendance'
    ws.append(['Subject', 'Date', 'Time'])
    for r in rows:
        ws.append(r)

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{student_name}_attendance.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')