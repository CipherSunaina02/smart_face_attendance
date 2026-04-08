from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

TIMETABLE_FILE = "timetable.xlsx"


def _ensure_timetable_file():
    if not os.path.exists(TIMETABLE_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Timetable"
        ws.append(["Day", "Start", "End", "Subject"])
        wb.save(TIMETABLE_FILE)


def get_current_subject():
    _ensure_timetable_file()
    wb = load_workbook(TIMETABLE_FILE)
    sheet = wb["Timetable"]

    today = datetime.now().strftime("%A")
    current_time = datetime.now().strftime("%H:%M")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        day, start, end, subject = row
        if day == today and start <= current_time <= end:
            return subject
    return None


def get_all_timetable():
    _ensure_timetable_file()
    wb = load_workbook(TIMETABLE_FILE)
    sheet = wb["Timetable"]
    entries = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        day, start, end, subject = row
        entries.append({"day": day or "", "start": start or "", "end": end or "", "subject": subject or ""})
    return entries


def save_timetable(entries):
    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"
    ws.append(["Day", "Start", "End", "Subject"])
    for e in entries:
        ws.append([e.get("day", ""), e.get("start", ""), e.get("end", ""), e.get("subject", "")])
    wb.save(TIMETABLE_FILE)
